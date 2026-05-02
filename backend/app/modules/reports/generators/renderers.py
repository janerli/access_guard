"""Report renderers — PDF (WeasyPrint), XLSX (openpyxl), CSV."""
import csv
import io
from typing import Any


class CsvRenderer:
    def render(self, data: dict) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        title = data.get("title", "")
        if title:
            writer.writerow([title])
            writer.writerow([])
        headers = data.get("headers", [])
        if headers:
            writer.writerow(headers)
        for row in data.get("rows", []):
            writer.writerow([str(c) for c in row])
        return buf.getvalue().encode("utf-8-sig")


class XlsxRenderer:
    def render(self, data: dict) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return self._fallback_csv(data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data.get("title", "Отчёт")[:31]

        row_idx = 1
        title = data.get("title", "")
        if title:
            ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True, size=14)
            row_idx += 2

        headers = data.get("headers", [])
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        if headers:
            row_idx += 1

        for row in data.get("rows", []):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            row_idx += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _fallback_csv(self, data: dict) -> bytes:
        return CsvRenderer().render(data)


class PdfRenderer:
    def render(self, data: dict) -> bytes:
        try:
            return self._render_weasyprint(data)
        except Exception:
            return self._render_plain(data)

    def _render_weasyprint(self, data: dict) -> bytes:
        from weasyprint import HTML
        html = self._build_html(data)
        return HTML(string=html).write_pdf()

    def _render_plain(self, data: dict) -> bytes:
        """Fallback: generate simple text-based content encoded as bytes."""
        lines = [f"# {data.get('title', 'Отчёт')}", ""]
        headers = data.get("headers", [])
        if headers:
            lines.append(" | ".join(str(h) for h in headers))
            lines.append("-" * 80)
        for row in data.get("rows", []):
            lines.append(" | ".join(str(c) for c in row))
        lines.append(f"\nИтого: {data.get('count', len(data.get('rows', [])))}")
        return "\n".join(lines).encode("utf-8")

    def _build_html(self, data: dict) -> str:
        title = data.get("title", "Отчёт")
        headers = data.get("headers", [])
        rows = data.get("rows", [])
        header_row = "".join(f"<th>{h}</th>" for h in headers)
        body_rows = "".join(
            "<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>"
            for row in rows
        )
        return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 10pt; padding: 20px; }}
  h1 {{ color: #2c3e50; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  th {{ background: #4472C4; color: white; padding: 6px 8px; text-align: left; }}
  td {{ padding: 4px 8px; border-bottom: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #f5f5f5; }}
  .count {{ margin-top: 12px; font-size: 9pt; color: #666; }}
</style></head>
<body>
<h1>{title}</h1>
<table><thead><tr>{header_row}</tr></thead><tbody>{body_rows}</tbody></table>
<p class="count">Строк: {data.get('count', len(rows))}</p>
</body></html>"""
